# -*- coding: utf-8 -*-
"""
Created on Wed Sep 11 13:06:38 2019

@author: Benny Daniel
"""


import openpyxl
import yaml as y

sourceFile = openpyxl.load_workbook('PATH_TO_FILE')
allSheetNames = sourceFile.sheetnames
dictionary={}
list_out=[]
keys=[]
values=[]
list4=[]
vlanslist=('vlanslist')
list4.append(vlanslist)
#print("All sheet names {} " .format(sourceFile.sheetnames))


def find_cell():
    for sheet in sourceFile.worksheets:
        max_row=sheet.max_row
        max_col=sheet.max_column
        for rows in range(1,max_row+1):
            for columns in range(1,max_col+1):  
                #cell_name = "{}{}".format(columns, rows)
                current_cell=sheet.cell(row=rows,column=columns)
                if current_cell.value == "VLAN ID": #Parameter(s) which define the header(s) of the table
                    Vid=current_cell #VLAN ID
                    description=Vid.offset(column=1) #Description
                    print("1:{} 2:{}".format(Vid.value,description.value))
                    while(Vid.offset(row=1).value != None):                 
                        Vid=Vid.offset(row=1)
                        description=description.offset(row=1)
                        keys.append(str(Vid.value))
                        values.append(description.value)
                    #print("cell position {} has value {}".format(cell_name, current_cell.value))
    zipped=zip(values,keys)
    for des,ids in zipped:
        dictionary={
                "name": des,
                "id": ids                
                }   
        list_out.append(dictionary)
    with open('Vlan_Data.yml','w') as output_file:
        y.dump(list4,output_file,default_flow_style=False)
    with open('Vlan_Data.yml','a') as output_file:
        y.dump(list_out,output_file,default_flow_style=False)
    with open('Vlan_Data.yml','r') as output_file:
        contents=output_file.read().split()
    print("\nSTRING\n", contents[0])
    print("List\n",list_out) 
    
        

find_cell()

    